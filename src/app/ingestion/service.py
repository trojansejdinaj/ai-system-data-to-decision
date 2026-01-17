from __future__ import annotations

import csv
import hashlib
import io
import json
import uuid
from dataclasses import dataclass
from datetime import UTC, datetime

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from app.db.models import IngestRun, RawRecord
from app.observability.logging import get_logger
from app.observability.run_tracking import RunTracker

REQUIRED_COLUMNS = ["source_id", "event_time", "value", "category"]


class IngestionError(ValueError):
    pass


@dataclass(frozen=True)
class IngestResult:
    run_id: uuid.UUID
    total_records: int
    inserted_records: int
    deduped_records: int
    per_file: dict[str, int]


def _parse_event_time(v: object) -> datetime:
    """Best-effort parse to tz-aware UTC datetime.

    We keep this intentionally strict-ish because the required schema contract
    guarantees an event_time column. If parsing fails, we raise so we don't
    silently write garbage keys that break dedupe.
    """
    if v is None:
        raise IngestionError("event_time is required")

    if isinstance(v, datetime):
        dt = v
    else:
        s = str(v).strip()
        if not s:
            raise IngestionError("event_time is required")
        # common ISO case: 2026-01-05T06:00:00Z
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        try:
            dt = datetime.fromisoformat(s)
        except ValueError as e:
            raise IngestionError(f"Invalid event_time format: {v!r}") from e

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    return dt.astimezone(UTC)


def _norm_str(v: object) -> str:
    return "" if v is None else str(v).strip()


def _extract_keys(payload: dict) -> tuple[str, datetime, str, str, str]:
    """Extract and normalize required keys + compute record hash.

    Returns: (source_id, event_time_dt, category, value, record_hash)
    """
    source_id = _norm_str(payload.get("source_id"))
    category = _norm_str(payload.get("category")).lower()
    value = _norm_str(payload.get("value"))
    event_dt = _parse_event_time(payload.get("event_time"))

    key = {
        "source_id": source_id,
        "event_time": event_dt.isoformat(),
        "category": category,
        "value": value,
    }
    encoded = json.dumps(key, sort_keys=True, separators=(",", ":")).encode("utf-8")
    record_hash = hashlib.sha256(encoded).hexdigest()
    return source_id, event_dt, category, value, record_hash


def _validate_headers(headers: list[str]) -> None:
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise IngestionError(f"Missing required columns: {missing}")


def _parse_csv(data: bytes) -> list[dict]:
    text = data.decode("utf-8")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    _validate_headers(headers)
    return [row for row in reader]


def _parse_xlsx(data: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() for h in rows[0]]
    _validate_headers(headers)
    out: list[dict] = []
    for r in rows[1:]:
        row = {headers[i]: r[i] for i in range(len(headers))}
        out.append(row)
    return out


def _parse_by_extension(filename: str, data: bytes) -> list[dict]:
    name = filename.lower()
    if name.endswith(".csv"):
        return _parse_csv(data)
    if name.endswith(".xlsx"):
        return _parse_xlsx(data)
    raise IngestionError(f"Unsupported file type: {filename}")


def ingest_files(db: Session, source: str, files: list[tuple[str, bytes]]) -> IngestResult:
    logger = get_logger(__name__)
    input_ref = ",".join([f[0] for f in files])
    tracker = RunTracker(db, logger, pipeline="ingest", input_ref=input_ref)

    run = IngestRun(source=source, files="\n".join([f[0] for f in files]), status="started")
    db.add(run)
    db.flush()  # get run.id

    per_file: dict[str, int] = {}
    total = 0
    inserted = 0
    deduped = 0

    try:
        for filename, data in files:
            with tracker.step("parse", meta={"filename": filename}):
                rows = _parse_by_extension(filename, data)
            per_file[filename] = len(rows)
            total += len(rows)

            if not rows:
                continue

            with tracker.step("upsert", meta={"filename": filename, "row_count": len(rows)}):
                now = datetime.now(UTC)
                values: list[dict] = []
                for idx, payload in enumerate(rows, start=1):
                    source_id, event_dt, category, value, record_hash = _extract_keys(payload)
                    values.append(
                        {
                            "id": uuid.uuid4(),
                            "run_id": run.id,
                            "row_num": idx,
                            "payload": payload,
                            "ingested_at": now,
                            "source": source,
                            "record_hash": record_hash,
                            "source_id": source_id,
                            "event_time": event_dt,
                            "category": category,
                            "value": value,
                        }
                    )

                stmt = pg_insert(RawRecord.__table__).values(values)
                stmt = stmt.on_conflict_do_nothing(index_elements=["source", "record_hash"])
                res = db.execute(stmt)
                added = int(res.rowcount or 0)
                inserted += added
                deduped += len(values) - added

        run.status = "success"
        db.commit()

        tracker.succeed()

        return IngestResult(
            run_id=run.id,
            total_records=total,
            inserted_records=inserted,
            deduped_records=deduped,
            per_file=per_file,
        )

    except Exception as e:
        tracker.fail(e)
        db.rollback()
        run.status = "failed"
        run.error = str(e)
        db.add(run)
        db.commit()
        raise
